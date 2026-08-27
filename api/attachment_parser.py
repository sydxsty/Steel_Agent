# -*- coding: utf-8 -*-
"""临时聊天附件解析器。

本模块只负责把单个聊天附件转换为 Markdown。它与 ``store_vectors.py``
完全隔离：不访问知识库 docs/markdown 目录、不生成向量，也不连接数据库。

后端通过独立 Python 子进程调用本文件。解析状态写入任务目录中的
``status.json``，因此即使 PDF OCR 耗时较长，前端也可以轮询进度；取消时
后端可以直接终止整个进程树，避免 MinerU 子任务残留。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Callable


MAX_OFFICE_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_PDF_PAGES = 300
MAX_IMAGE_PIXELS = 50_000_000


class AttachmentParseError(RuntimeError):
    """附件无法安全、完整地转换为 Markdown。"""


def _atomic_write_text(path: Path, content: str) -> None:
    """先写临时文件再替换，避免状态轮询读到半个 JSON。"""
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(content, encoding="utf-8")
    os.replace(temporary, path)


def _write_status(task_dir: Path, status: str, progress: int, message: str = "") -> None:
    """更新供主进程和前端读取的解析状态。"""
    payload = {
        "status": status,
        "progress": max(0, min(100, int(progress))),
        "message": str(message or ""),
    }
    _atomic_write_text(
        task_dir / "status.json",
        json.dumps(payload, ensure_ascii=False),
    )


def _escape_markdown_cell(value) -> str:
    """清理 Markdown 表格单元格中的换行和竖线。"""
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").replace("|", "\\|").strip()


def _rows_to_markdown(rows: list[list]) -> str:
    """把二维数据转为列数稳定的 Markdown 表格。"""
    clean_rows = [
        [_escape_markdown_cell(cell) for cell in row]
        for row in rows
        if any(cell not in (None, "") for cell in row)
    ]
    if not clean_rows:
        return ""
    column_count = max(len(row) for row in clean_rows)
    clean_rows = [row + [""] * (column_count - len(row)) for row in clean_rows]
    lines = [
        "| " + " | ".join(clean_rows[0]) + " |",
        "| " + " | ".join(["---"] * column_count) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in clean_rows[1:])
    return "\n".join(lines)


def _validate_office_archive(path: Path, expected_folder: str) -> None:
    """校验 DOCX/XLSX ZIP 结构并限制解压后体积，防止压缩炸弹。"""
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            if "[Content_Types].xml" not in names or not any(
                name.startswith(expected_folder + "/") for name in names
            ):
                raise AttachmentParseError("文件内容与扩展名不匹配")
            expanded_size = sum(item.file_size for item in archive.infolist())
            if expanded_size > MAX_OFFICE_UNCOMPRESSED_BYTES:
                raise AttachmentParseError("Office 文件解压后体积超过安全上限")
    except zipfile.BadZipFile as exc:
        raise AttachmentParseError("Office 文件已损坏或不是有效的现代 Office 格式") from exc


def _parse_text(path: Path) -> str:
    """按常见中文文本编码读取 TXT/Markdown。"""
    raw = path.read_bytes()
    if b"\x00" in raw:
        raise AttachmentParseError("文本文件包含二进制内容")
    for encoding in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            text = raw.decode(encoding)
            if text.strip():
                return text.strip()
        except UnicodeDecodeError:
            continue
    raise AttachmentParseError("无法识别文本文件编码或文件内容为空")


def _parse_docx(path: Path) -> str:
    """按 Word 正文顺序保留段落、标题和表格。"""
    _validate_office_archive(path, "word")
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    document = Document(str(path))
    parts: list[str] = []
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            paragraph = Paragraph(child, document)
            text = paragraph.text.strip()
            if not text:
                continue
            style_name = str(getattr(paragraph.style, "name", "") or "")
            match = re.search(r"(?:Heading|标题)\s*([1-6])", style_name, re.I)
            parts.append(("#" * int(match.group(1)) + " " + text) if match else text)
        elif child.tag.endswith("}tbl"):
            table = Table(child, document)
            markdown = _rows_to_markdown([
                [cell.text for cell in row.cells]
                for row in table.rows
            ])
            if markdown:
                parts.append(markdown)
    result = "\n\n".join(parts).strip()
    if not result:
        raise AttachmentParseError("Word 文件没有可提取的正文或表格")
    return result


def _parse_xlsx(path: Path) -> str:
    """逐工作表读取 Excel；公式保留为公式文本而不是陈旧缓存值。"""
    _validate_office_archive(path, "xl")
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=False)
    parts: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            markdown = _rows_to_markdown(rows)
            if markdown:
                parts.append(f"## 工作表：{worksheet.title}\n\n{markdown}")
    finally:
        workbook.close()
    result = "\n\n".join(parts).strip()
    if not result:
        raise AttachmentParseError("Excel 文件没有可提取的非空单元格")
    return result


def _validate_pdf(path: Path) -> int:
    """验证 PDF 签名、加密状态和页数。"""
    if not path.read_bytes()[:5] == b"%PDF-":
        raise AttachmentParseError("文件内容与 PDF 扩展名不匹配")
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    if reader.is_encrypted:
        raise AttachmentParseError("暂不支持加密 PDF")
    page_count = len(reader.pages)
    if page_count <= 0:
        raise AttachmentParseError("PDF 没有有效页面")
    if page_count > MAX_PDF_PAGES:
        raise AttachmentParseError(f"PDF 超过 {MAX_PDF_PAGES} 页安全上限")
    return page_count


def _parse_pdf_with_mineru(path: Path, task_dir: Path) -> str:
    """使用项目已有 MinerU 做版面分析、表格识别和扫描件 OCR。"""
    models_dir = Path.home() / "magic-pdf-models"
    if not models_dir.exists():
        return ""
    try:
        from magic_pdf.tools.common import do_parse, prepare_env
    except ImportError:
        return ""

    output_dir = task_dir / "mineru_output"
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        prepare_env(str(output_dir), path.stem, "auto")
        do_parse(
            output_dir=str(output_dir),
            pdf_file_name=path.stem,
            pdf_bytes_or_dataset=path.read_bytes(),
            model_list=[],
            parse_method="auto",
            lang="ch",
            f_dump_md=True,
            f_dump_middle_json=False,
            f_dump_model_json=False,
            f_dump_orig_pdf=False,
            f_dump_content_list=True,
            table_enable=True,
            formula_enable=False,
        )
        markdown_files = list(output_dir.rglob(f"{path.stem}.md"))
        if markdown_files:
            return markdown_files[0].read_text(encoding="utf-8").strip()
    except Exception:
        # MinerU 失败属于可回退错误，后续继续使用文本型 PDF 解析器。
        return ""
    return ""


def _parse_pdf_with_plumber(path: Path) -> str:
    """使用 pdfplumber 按页提取正文和表格。"""
    try:
        import pdfplumber
    except ImportError:
        return ""
    parts: list[str] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                page_parts: list[str] = []
                text = (page.extract_text() or "").strip()
                if text:
                    page_parts.append(text)
                for table in page.extract_tables() or []:
                    markdown = _rows_to_markdown(table or [])
                    if markdown:
                        page_parts.append(markdown)
                if page_parts:
                    parts.append(f"## 第 {page_number} 页\n\n" + "\n\n".join(page_parts))
    except Exception:
        return ""
    return "\n\n".join(parts).strip()


def _parse_pdf_with_fitz(path: Path) -> str:
    """使用 PyMuPDF 作为文本提取回退。"""
    try:
        import fitz
    except ImportError:
        return ""
    parts: list[str] = []
    try:
        document = fitz.open(str(path))
        for page_number, page in enumerate(document, start=1):
            text = page.get_text("text").strip()
            if text:
                parts.append(f"## 第 {page_number} 页\n\n{text}")
        document.close()
    except Exception:
        return ""
    return "\n\n".join(parts).strip()


def _parse_pdf_with_pypdf(path: Path) -> str:
    """使用 PyPDF 作为最后的纯文本回退。"""
    from pypdf import PdfReader

    parts: list[str] = []
    reader = PdfReader(str(path))
    for page_number, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            parts.append(f"## 第 {page_number} 页\n\n{text}")
    return "\n\n".join(parts).strip()


def _parse_pdf(path: Path, task_dir: Path, progress: Callable[[int, str], None]) -> str:
    """按 MinerU → pdfplumber → PyMuPDF → PyPDF 的顺序解析 PDF。"""
    _validate_pdf(path)
    progress(35, "正在执行 PDF 版面分析与文字识别")
    for parser in (
        lambda: _parse_pdf_with_mineru(path, task_dir),
        lambda: _parse_pdf_with_plumber(path),
        lambda: _parse_pdf_with_fitz(path),
        lambda: _parse_pdf_with_pypdf(path),
    ):
        content = parser()
        if content.strip():
            return content.strip()
    raise AttachmentParseError("PDF 未提取到可用文字；扫描件 OCR 也未返回内容")


def _parse_image(path: Path, task_dir: Path, progress: Callable[[int, str], None]) -> str:
    """校验图片后转为单页 PDF，复用 MinerU OCR，避免依赖 Tesseract。"""
    from PIL import Image, UnidentifiedImageError

    try:
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            if image.width * image.height > MAX_IMAGE_PIXELS:
                raise AttachmentParseError("图片像素总量超过安全上限")
            converted = image.convert("RGB")
            pdf_path = task_dir / "image_ocr.pdf"
            converted.save(pdf_path, "PDF", resolution=150.0)
    except UnidentifiedImageError as exc:
        raise AttachmentParseError("文件内容与图片扩展名不匹配或图片已损坏") from exc
    progress(35, "正在识别图片文字")
    content = _parse_pdf_with_mineru(pdf_path, task_dir)
    if not content:
        # 文本回退对图片型 PDF 通常为空，但保留统一处理逻辑。
        content = _parse_pdf_with_fitz(pdf_path)
    if not content.strip():
        raise AttachmentParseError("图片中未识别到可用文字")
    return content.strip()


def parse_attachment(task_dir: Path) -> Path:
    """读取任务元数据，将附件转换为完整 Markdown 并返回结果文件路径。"""
    metadata_file = task_dir / "metadata.json"
    if not metadata_file.exists():
        raise AttachmentParseError("附件任务元数据不存在")
    metadata = json.loads(metadata_file.read_text(encoding="utf-8"))
    source_path = task_dir / "source" / metadata["stored_name"]
    original_name = str(metadata["original_name"])
    extension = source_path.suffix.lower()

    def report(value: int, message: str) -> None:
        _write_status(task_dir, "parsing", value, message)

    report(15, "正在校验附件")
    if extension in {".txt", ".md", ".markdown"}:
        body = _parse_text(source_path)
    elif extension == ".docx":
        report(35, "正在提取 Word 段落和表格")
        body = _parse_docx(source_path)
    elif extension == ".xlsx":
        report(35, "正在提取 Excel 工作表")
        body = _parse_xlsx(source_path)
    elif extension == ".pdf":
        body = _parse_pdf(source_path, task_dir, report)
    elif extension in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}:
        body = _parse_image(source_path, task_dir, report)
    else:
        raise AttachmentParseError(f"不支持的附件格式：{extension or '无扩展名'}")

    report(90, "正在整理 Markdown")
    result = f"# 附件：{original_name}\n\n{body.strip()}\n"
    output_path = task_dir / "attachment.md"
    _atomic_write_text(output_path, result)
    _write_status(task_dir, "ready", 100, "解析完成")
    return output_path


def main() -> int:
    """子进程命令行入口；失败信息统一写回 status.json。"""
    parser = argparse.ArgumentParser(description="解析临时聊天附件")
    parser.add_argument("--task-dir", required=True)
    args = parser.parse_args()
    task_dir = Path(args.task_dir).resolve()
    try:
        parse_attachment(task_dir)
        return 0
    except Exception as exc:
        _write_status(task_dir, "error", 100, str(exc))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
